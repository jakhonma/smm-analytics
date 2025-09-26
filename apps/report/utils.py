import io
import requests
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter


def download_kpi_report(data: list):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employee KPI"

    headers = ["Xodim", "Avatar", "Kanal", "Ijtimoiy tarmoqlar", "Ball", "Jami ball", "KPI"]
    ws.append(headers)

    current_row = 2

    for emp in data:
        # 🔹 Kanal bo‘yicha guruhlash
        grouped = defaultdict(list)
        for d in emp["data"]:
            grouped[d["channel"]].append(d)

        rows_count = sum(len(v) for v in grouped.values())
        start_row = current_row
        end_row = current_row + rows_count - 1

        for channel, items in grouped.items():
            ch_start = current_row
            ch_end = current_row + len(items) - 1

            # Social tarmoqlarni yozish
            for d in items:
                ws.cell(row=current_row, column=4, value=d["social_network"])
                ws.cell(row=current_row, column=5, value=d["score"])
                current_row += 1

            # Kanalni merge qilish
            ws.merge_cells(start_row=ch_start, start_column=3, end_row=ch_end, end_column=3)
            ws.cell(row=ch_start, column=3, value=channel)

        # Employee, Avatar, Total Score, KPI ni merge qilish
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
        ws.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)

        ws.cell(row=start_row, column=1, value=emp["employee"])
        ws.cell(row=start_row, column=6, value=emp["total_score"])
        ws.cell(row=start_row, column=7, value=emp["kpi"])

        # Avatar rasmi
        if emp["avatar"]:
            try:
                img_data = requests.get(emp["avatar"], stream=True).content
                img_stream = io.BytesIO(img_data)
                img = XLImage(img_stream)
                img.width, img.height = 50, 50
                ws.add_image(img, f"{get_column_letter(2)}{start_row}")
            except Exception:
                ws.cell(row=start_row, column=2, value="(Avatar yuklanmadi)")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
